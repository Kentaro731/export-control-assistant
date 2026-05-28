import os
import io
import json
from datetime import datetime
from flask import (
    Flask, render_template, jsonify, request, Response, send_file,
)
import anthropic
from openpyxl import Workbook, load_workbook
from export_law_data import (
    is_white_country, is_designated_country,
    has_danger_keywords, has_military_end_use, get_hs_hint,
    CONTROLLED_CATEGORIES_1_TO_15,
)

app = Flask(__name__)
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

# === Basic認証（社内ポータルと共通の資格情報で統一）===
# 資格情報は環境変数 BASIC_AUTH_USER / BASIC_AUTH_PASS で設定する。
# ※本リポジトリは公開のため、パスワードはコードに書かず Render の環境変数にのみ置く。
#   社内ポータル（shinx-internal-apps）と同じ値を設定すれば、社員は1組の資格情報で両方を利用できる。
BASIC_AUTH_USER = os.environ.get("BASIC_AUTH_USER", "shinx")
BASIC_AUTH_PASS = os.environ.get("BASIC_AUTH_PASS", "")  # 秘密はコードに残さない


def _request_auth():
    """401を返してブラウザに認証ダイアログを表示させる"""
    return Response(
        "認証が必要です。正しいユーザー名とパスワードを入力してください。",
        status=401,
        headers={"WWW-Authenticate": 'Basic realm="Shinx Internal Apps"'},
    )


@app.before_request
def enforce_basic_auth():
    # ヘルスチェックは認証不要
    if request.path == "/healthz":
        return None
    # パスワード未設定なら安全側に全拒否（誤って開放しないため）
    if not BASIC_AUTH_PASS:
        return Response(
            "サーバー設定エラー：BASIC_AUTH_PASS が未設定です。管理者に連絡してください。",
            status=503,
        )
    auth = request.authorization
    if not auth or auth.username != BASIC_AUTH_USER or auth.password != BASIC_AUTH_PASS:
        return _request_auth()
    return None


HISTORY_FILE = os.path.join(os.path.dirname(__file__), "judgment_history.json")

SYSTEM_PROMPT = """あなたは日本の安全保障輸出管理（外為法・輸出貿易管理令）の専門家AIです。
非鉄金属卸（アルミ・銅・真鍮・チタン等の材料プレート・切り板を扱う企業）の
輸出担当者を支援します。

## 最重要の前提（位置づけ）
- あなたの出力は「最終的な該非判定」ではなく、担当者が確認するための【暫定スクリーニング（一次確認の補助）】である。
- 最終的な該非判定は、必ず人間（輸出管理担当者）が公式の「項目別対比表」で確定する。あなたは断定せず、確認すべき論点・根拠・不明点を提示する。
- 特にリスト規制（輸出令別表第1の1〜15項）の該当性は、貨物の仕様値と省令のしきい値を対比表で突合して確定するもの。あなたは数値しきい値を断定せず、「対比表での突合が必要」と促すこと。情報が不足する場合は推測で「非該当」と言い切らない。

## 判定フロー（5ステップ）

STEP1: 輸出令別表第1の1〜15項への該当確認
- 1〜15項は主に武器・核・化学・生物兵器、軍用装備品等
- 金属材料・プレート・切り板は通常「非該当」だが、用途・仕様によっては確認要
- 高強度・特殊合金で軍事用途が疑われる場合は要注意
- 【重要】用途・商流・顧客に軍事・防衛・兵器関連の用途（軍用／軍事／防衛省／自衛隊／戦闘機・軍用機／戦車・装甲車／潜水艦・軍艦／ミサイル・弾道／砲・銃・弾薬等）が明示・示唆されている場合は、別表第1（特に9項航空機用エンジン・10項航空機宇宙機器・11項軍用艦船・12項戦車装甲車両・13〜14項軍用電子/誘導装置）への該当可能性が高い。この場合STEP1の結果を「非該当」にしてはならない。最低でも「要確認」、用途が明確な軍事用途であれば「該当」とすること。

STEP2: 16項（キャッチオール規制）への該当確認
- 輸出令別表第1の1〜15項非該当でも、大量破壊兵器等の開発・製造に使われるおそれがある場合は規制対象
- 仕向け国・用途・顧客の総合判断が必要

STEP3: 仕向地が輸出令別表第3（ホワイト国）か確認
- ホワイト国へのSTEP1・2非該当品 → 輸出許可不要（OK）
- 非ホワイト国 → STEP4へ
- 【重要】ホワイト国向けであっても、別表第1（1〜15項）に該当する貨物（リスト規制品）は輸出許可が必要。ホワイト国であることはリスト規制（STEP1）の免除理由にはならない。免除されるのはキャッチオール規制（STEP2・4）のみ。

STEP4: キャッチオール規制の用途確認（非ホワイト国向けの場合のみ）
①大量破壊兵器等の開発・製造・使用・貯蔵の用途かどうか
②おそれ省令別表に掲げる行為（輸出令※5）に当たるかどうか
③仕向地が輸出令別表第3の2（北朝鮮・イラン等）かつ通常兵器への用途かどうか
- ホワイト国向けでSTEP1が「非該当」の場合のみ「スキップ」可。STEP1が「該当」または「要確認」の場合は、ホワイト国でもスキップせず用途を精査すること。

STEP5: STEP4に該当する場合、用途・仕向地・エンドユーザーで個別判断

## 総合判断（overall）の整合性ルール（厳守）
- STEP1が「該当」、またはSTEP1〜4のいずれかが「該当」 → overallは「要許可」
- STEP1〜4のいずれかが「要確認」 → overallは最低でも「要確認」（「OK」にしてはならない）
- 明確な軍事用途が示されている場合 → 安全側に倒し「要許可」とすること
- 全ステップが「非該当／ホワイト国／問題なし」のときのみ overallを「OK」にできる

## 出力形式
必ず以下のJSON形式のみで回答してください。説明文は不要。

各stepには必ず "confidence"（"高"／"中"／"低"）を付けること。情報不足・推測を含む場合は「低」または「中」とする。
"unverifiable" には、AIだけでは確認できず担当者が必ず確認すべき事項（例：エンドユーザーの実在・素性、最終用途の真偽、製品仕様値の対比表突合、再輸出の有無など）を配列で列挙すること。

{
  "overall": "OK" または "要確認" または "要許可",
  "steps": [
    {"step": 1, "title": "別表第1（1〜15項）該当確認", "result": "非該当" または "要確認" または "該当", "reason": "判断根拠を簡潔に", "confidence": "高" または "中" または "低"},
    {"step": 2, "title": "16項（キャッチオール）確認", "result": "非該当" または "要確認" または "該当", "reason": "判断根拠を簡潔に", "confidence": "高" または "中" または "低"},
    {"step": 3, "title": "仕向地（別表第3）確認", "result": "ホワイト国" または "非ホワイト国" または "指定国", "reason": "判断根拠を簡潔に", "confidence": "高" または "中" または "低"},
    {"step": 4, "title": "用途確認（キャッチオール）", "result": "非該当" または "要確認" または "該当" または "スキップ", "reason": "判断根拠または「ホワイト国のためスキップ」", "confidence": "高" または "中" または "低"},
    {"step": 5, "title": "総合判断", "result": "問題なし" または "要精査" または "輸出不可", "reason": "最終的な根拠", "confidence": "高" または "中" または "低"}
  ],
  "unverifiable": ["担当者が必ず確認すべき事項を列挙"],
  "recommendation": "担当者への具体的な推奨アクション（1〜3文）",
  "hs_hint": "推定HSコード（参考）",
  "caution": "注意事項や免責文言"
}"""


def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_history(entry):
    history = load_history()
    history.insert(0, entry)
    history = history[:100]  # 最新100件のみ保持
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


@app.route("/healthz")
def healthz():
    """Renderヘルスチェック用（認証不要）"""
    return {"status": "ok"}, 200


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/judge", methods=["POST"])
def judge():
    data = request.json
    customer = data.get("customer", "")
    material = data.get("material", "")
    dimensions = data.get("dimensions", "")
    purpose = data.get("purpose", "")
    trade_flow = data.get("trade_flow", "")
    destination = data.get("destination", "")

    # 事前チェック（ルールベース補助）
    is_white = is_white_country(destination)
    is_designated = is_designated_country(destination)
    scan_text = purpose + " " + trade_flow + " " + customer + " " + material
    danger_kws = has_danger_keywords(scan_text)
    military_kws = has_military_end_use(scan_text)
    hs_hint = get_hs_hint(material)

    user_message = f"""以下の輸出貨物について、5ステップの該非判定を行ってください。

【顧客名】{customer}
【材質】{material}
【寸法】{dimensions}
【用途】{purpose}
【商流（中間業者等）】{trade_flow}
【仕向け国】{destination}

【補助情報（システム事前チェック）】
- 別表第3ホワイト国: {"YES" if is_white else "NO"}
- 別表第3の2指定国: {"YES" if is_designated else "NO"}
- 大量破壊兵器関連の危険キーワード検出: {", ".join(danger_kws) if danger_kws else "なし"}
- ⚠️通常兵器・軍事用途キーワード検出: {", ".join(military_kws) if military_kws else "なし"}
{"  → 【警告】軍事用途キーワードを検出。ホワイト国向けでも別表第1（リスト規制）該当性を厳格に評価し、STEP1を非該当にせず、overallは要許可（または最低でも要確認）とすること。" if military_kws else ""}
- 推定HSコード: {hs_hint}
- 輸出令別表第1（1〜15項）主要カテゴリ:
{chr(10).join(CONTROLLED_CATEGORIES_1_TO_15)}

JSON形式のみで回答してください。"""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_message}],
        )
        raw = response.content[0].text.strip()
        # JSON部分のみ抽出
        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()

        result = json.loads(raw)
        result["hs_hint"] = result.get("hs_hint") or hs_hint

        # 安全ネット（ガードレール）：軍事用途キーワードを検出しているのに
        # AIが「OK」と返した場合は、安全側に倒して最低「要確認」へ強制的に引き上げる。
        if military_kws and result.get("overall") == "OK":
            result["overall"] = "要確認"
            warn = f"※システム自動補正：軍事用途キーワード（{', '.join(military_kws)}）を検出したため、判定を『要確認』に引き上げました。別表第1（リスト規制）該当性を担当者が必ず確認してください。"
            result["recommendation"] = warn + " " + result.get("recommendation", "")

        entry = {
            "id": datetime.now().strftime("%Y%m%d%H%M%S"),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "customer": customer,
            "material": material,
            "destination": destination,
            "purpose": purpose,
            "overall": result.get("overall", "要確認"),
            "recommendation": result.get("recommendation", ""),
            # 人間の承認ワークフロー（初期は未確定）
            "status": "未確定",
            "approver": "",
            "approved_at": "",
            "final_decision": "",
            "taikohyo_checked": False,
        }
        save_history(entry)

        return jsonify({"status": "ok", "result": result, "entry_id": entry["id"]})

    except json.JSONDecodeError:
        return jsonify({"status": "ok", "result": {
            "overall": "要確認",
            "steps": [
                {"step": i, "title": t, "result": "確認中", "reason": "AI応答の解析に失敗しました。再度お試しください。"}
                for i, t in enumerate([
                    "別表第1（1〜15項）該当確認",
                    "16項（キャッチオール）確認",
                    "仕向地（別表第3）確認",
                    "用途確認（キャッチオール）",
                    "総合判断",
                ], 1)
            ],
            "recommendation": "AI応答の解析に失敗しました。入力内容を確認の上、再度お試しください。",
            "hs_hint": hs_hint,
            "caution": "本ツールはAI補助判定です。最終判断は担当者が行ってください。",
        }})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/history", methods=["GET"])
def get_history():
    return jsonify(load_history())


@app.route("/api/history/<entry_id>", methods=["DELETE"])
def delete_history(entry_id):
    history = load_history()
    history = [h for h in history if h.get("id") != entry_id]
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    return jsonify({"status": "ok"})


@app.route("/api/history/<entry_id>/approve", methods=["POST"])
def approve_history(entry_id):
    """担当者による最終判断の記録（人間の承認ワークフロー）"""
    data = request.json or {}
    approver = (data.get("approver") or "").strip()
    final_decision = (data.get("final_decision") or "").strip()
    taikohyo_checked = bool(data.get("taikohyo_checked"))
    if not approver or not final_decision:
        return jsonify({"status": "error",
                        "message": "確認者名と最終判断は必須です。"}), 400
    history = load_history()
    found = False
    for h in history:
        if h.get("id") == entry_id:
            h["status"] = "承認済み"
            h["approver"] = approver
            h["approved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            h["final_decision"] = final_decision
            h["taikohyo_checked"] = taikohyo_checked
            found = True
            break
    if not found:
        return jsonify({"status": "error", "message": "該当履歴が見つかりません。"}), 404
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    return jsonify({"status": "ok"})


# =====================================================================
# Excel 取り込み機能（単票読込み・一括判定の入口）
# =====================================================================

# 入力フィールドの正規名と、Excel側のヘッダー候補（柔軟マッチング用）
EXCEL_FIELD_ALIASES = {
    "customer":    ["顧客名", "顧客", "客先", "得意先", "取引先", "社名", "会社名"],
    "material":    ["材質", "品名", "材料", "品目", "材料名"],
    "dimensions":  ["寸法", "サイズ", "規格", "サイズ・寸法"],
    "purpose":     ["用途", "最終用途", "用途・最終用途", "使用用途"],
    "trade_flow":  ["商流", "取引フロー", "中間業者", "流通経路", "ルート"],
    "destination": ["仕向け国", "仕向地", "仕向先国", "輸出先", "輸出国", "送り先国", "国"],
}

EXCEL_FIELD_LABEL = {
    "customer": "顧客名",
    "material": "材質",
    "dimensions": "寸法",
    "purpose": "用途",
    "trade_flow": "商流",
    "destination": "仕向け国",
}


def _normalize_header(s):
    if s is None:
        return ""
    return str(s).strip().replace("　", "").replace(" ", "").replace("\n", "")


def _match_header_to_field(header):
    """Excelヘッダー文字列を正規フィールド名にマッピング（部分一致を含む柔軟マッチング）"""
    if not header:
        return None
    h = _normalize_header(header)
    if not h:
        return None
    for field, aliases in EXCEL_FIELD_ALIASES.items():
        for a in aliases:
            an = _normalize_header(a)
            if h == an or an in h or h in an:
                return field
    return None


@app.route("/api/excel/template", methods=["GET"])
def excel_template():
    """貨物情報入力テンプレートの.xlsxをその場で生成して返す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "貨物情報"
    headers = ["顧客名", "材質", "寸法", "用途", "商流", "仕向け国"]
    ws.append(headers)
    # サンプル行（参考用・削除して使ってください）
    ws.append([
        "（例）株式会社〇〇製作所",
        "アルミ A6061",
        "200×300×10mm",
        "自動車部品の試作用ブラケット材料",
        "当社 → 商社A → 米国メーカー",
        "アメリカ",
    ])
    # 列幅を整える
    widths = [22, 22, 22, 32, 32, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="貨物情報テンプレート.xlsx",
    )


MAX_BATCH_ROWS = 20  # 一括判定の上限（コスト・タイムアウト保護）


@app.route("/api/excel/parse", methods=["POST"])
def excel_parse():
    """アップロードされたExcelを解析し、貨物情報の行リストを返す（判定は別途）"""
    f = request.files.get("file")
    if not f:
        return jsonify({"status": "error", "message": "ファイルが添付されていません。"}), 400
    name = (f.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        return jsonify({"status": "error", "message": "Excel(.xlsx)ファイルをアップロードしてください。"}), 400
    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Excel読み込みに失敗しました: {e}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"status": "error", "message": "シートが空です。"}), 400

    # ヘッダー行を検出（最初の非空行）
    header_row_idx = -1
    headers = []
    for i, r in enumerate(rows):
        if r and any(c is not None and str(c).strip() != "" for c in r):
            header_row_idx = i
            headers = list(r)
            break
    if header_row_idx < 0:
        return jsonify({"status": "error", "message": "ヘッダー行が見つかりません。"}), 400

    # 列インデックス→フィールドのマッピング
    col_to_field = {}
    for ci, h in enumerate(headers):
        field = _match_header_to_field(h)
        if field and field not in col_to_field.values():
            col_to_field[ci] = field

    if not col_to_field:
        return jsonify({
            "status": "error",
            "message": "ヘッダーに認識可能な列が見つかりませんでした。テンプレートを参考に列名を設定してください（顧客名・材質・寸法・用途・商流・仕向け国）。"
        }), 400

    # データ行を抽出
    parsed = []
    skipped = 0
    sample_marker = "（例）"
    for r in rows[header_row_idx + 1:]:
        if not r:
            continue
        item = {"customer": "", "material": "", "dimensions": "",
                "purpose": "", "trade_flow": "", "destination": ""}
        nonempty = False
        for ci, field in col_to_field.items():
            if ci < len(r) and r[ci] is not None:
                v = str(r[ci]).strip()
                if v:
                    item[field] = v
                    nonempty = True
        if not nonempty:
            continue
        # サンプル行（テンプレの「（例）」始まり）は除外
        if item.get("customer", "").startswith(sample_marker):
            skipped += 1
            continue
        # 必須最低限：材質と仕向け国がないとスキップ
        if not item["material"] or not item["destination"]:
            skipped += 1
            continue
        parsed.append(item)
        if len(parsed) >= MAX_BATCH_ROWS:
            break

    matched_fields = [EXCEL_FIELD_LABEL[v] for v in set(col_to_field.values())]
    return jsonify({
        "status": "ok",
        "rows": parsed,
        "skipped": skipped,
        "matched_fields": matched_fields,
        "max_batch_rows": MAX_BATCH_ROWS,
    })


# =====================================================================
# 参照資料の最新確認（判定フロー内で言及している公式ページ）
# =====================================================================
import urllib.request as _urllib_req
import urllib.error as _urllib_err

DOC_REFS = [
    {
        "id": "matrix",
        "name": "項目別対比表",
        "url": "https://www.meti.go.jp/policy/anpo/matrix_index.html",
        "desc": "リスト規制（STEP1）判定の基準。製品仕様を突合する公式資料",
    },
    {
        "id": "anpo_top",
        "name": "安全保障輸出管理（経産省トップ）",
        "url": "https://www.meti.go.jp/policy/anpo/index.html",
        "desc": "法令・規制・通知の最新情報。まず確認するページ",
    },
    {
        "id": "whitelist",
        "name": "ホワイト国（別表第3）一覧",
        "url": "https://www.meti.go.jp/policy/anpo/law01.html",
        "desc": "キャッチオール通常兵器免除対象の42カ国リスト（STEP3）",
    },
    {
        "id": "catchall",
        "name": "キャッチオール規制",
        "url": "https://www.meti.go.jp/policy/anpo/catch_all.html",
        "desc": "大量破壊兵器・通常兵器の用途規制（STEP2）の詳細基準",
    },
    {
        "id": "egov_law",
        "name": "輸出貿易管理令（e-Gov法令）",
        "url": "https://elaws.e-gov.go.jp/document?lawid=374CO0000000378",
        "desc": "別表第1〜3の法令原文（最新の改正内容を確認）",
    },
]


def _check_one_doc(doc):
    """1件の参照資料にアクセスし、到達可否と最終更新日を取得する"""
    result = dict(doc)
    try:
        req = _urllib_req.Request(
            doc["url"],
            headers={"User-Agent": "Mozilla/5.0 (compatible; ExportCheckTool/1.0)"},
        )
        resp = _urllib_req.urlopen(req, timeout=12)
        result["http_status"] = resp.status
        result["last_modified"] = resp.headers.get("Last-Modified") or "—"
        result["reachable"] = True
        result["error"] = None
    except _urllib_err.HTTPError as e:
        result["http_status"] = e.code
        result["last_modified"] = "—"
        result["reachable"] = False
        result["error"] = f"HTTP {e.code}"
    except Exception as e:
        result["http_status"] = "—"
        result["last_modified"] = "—"
        result["reachable"] = False
        result["error"] = str(e)[:60]
    return result


@app.route("/api/check_docs")
def check_docs():
    """参照資料（経産省・e-Gov）のアクセス確認エンドポイント（5件を並列取得）"""
    from concurrent.futures import ThreadPoolExecutor
    import datetime as _dt

    jst = _dt.timezone(_dt.timedelta(hours=9))
    checked_at = _dt.datetime.now(jst).strftime("%Y-%m-%d %H:%M JST")

    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(_check_one_doc, doc): doc for doc in DOC_REFS}
        results_map = {}
        for fut in futures:
            doc = futures[fut]
            try:
                r = fut.result(timeout=15)
            except Exception as e:
                r = dict(doc, http_status="—", last_modified="—",
                         reachable=False, error=str(e)[:60])
            r["checked_at"] = checked_at
            results_map[doc["id"]] = r

    results = [results_map[d["id"]] for d in DOC_REFS if d["id"] in results_map]
    return jsonify({"status": "ok", "results": results, "checked_at": checked_at})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5002))
    app.run(debug=False, host="0.0.0.0", port=port)
