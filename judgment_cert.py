"""
該非判定書 生成モジュール

generate_xls(entry) -> (bytes, error_str)
    テンプレートあり → xlrd + xlutils.copy でシンクス .xls を生成（板/丸棒・材質別）
    テンプレートなし → openpyxl で .xlsx を生成（フォールバック）

generate_pdf(entry) -> (bytes, error_str)
    reportlab で A4 PDF 証明書を生成して返す。

テンプレート取得優先順位:
  1. Render Secret File: /etc/secrets/shinx_template_b64.txt（base64テキスト）
  2. 環境変数 SHINX_TEMPLATE_XLS_BASE64（base64, 予備）
  3. OneDriveローカルパス（開発時）
"""

import os, io, base64
from datetime import date

# ── xlwt Python3.12 互換パッチ（None フォーマット文字列対策）──
try:
    import xlwt.UnicodeUtils as _xlu
    import xlwt.BIFFRecords as _xlb
    if not getattr(_xlu, '_safe_patched', False):
        _orig_upack2 = _xlu.upack2
        def _safe_upack2(s, encoding='ascii'):
            if s is None:
                s = ''
            return _orig_upack2(s, encoding)
        _xlu.upack2 = _safe_upack2
        _xlb.upack2 = _safe_upack2
        _xlu._safe_patched = True
except Exception:
    pass

# ── テンプレートキャッシュ ──────────────────────────────────────────
_TEMPLATE_BYTES: bytes | None = None

# Render Secret File パス（base64テキスト形式で保存）
_SECRET_FILE = "/etc/secrets/shinx_template_b64.txt"

_TEMPLATE_LOCAL = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - 柏原 健太郎",
    "★2. シンクスコーポレーション",
    "★ShinxAI", "❷輸出判定ツール", "該非判定書　原本.xls",
)


def _get_template() -> bytes | None:
    global _TEMPLATE_BYTES
    if _TEMPLATE_BYTES:
        return _TEMPLATE_BYTES
    # 優先1: Render Secret File（base64テキスト）
    if os.path.exists(_SECRET_FILE):
        with open(_SECRET_FILE, "r") as f:
            b64 = f.read().strip()
        if b64:
            _TEMPLATE_BYTES = base64.b64decode(b64)
            return _TEMPLATE_BYTES
    # 優先2: 環境変数（予備）
    b64 = os.environ.get("SHINX_TEMPLATE_XLS_BASE64", "")
    if b64:
        _TEMPLATE_BYTES = base64.b64decode(b64)
        return _TEMPLATE_BYTES
    # 優先3: ローカルファイル（開発時）
    if os.path.exists(_TEMPLATE_LOCAL):
        with open(_TEMPLATE_LOCAL, "rb") as f:
            _TEMPLATE_BYTES = f.read()
        return _TEMPLATE_BYTES
    return None


# ── 材質→シート名マッピング（長い（より具体的な）キーワードを先に検索）──
_MATERIAL_RULES = [
    ("A2017B",  "A2017B 75φ以下"),
    ("A7075B",  "A7075B 75φ以下"),
    ("A5052",   "A5052"),
    ("A2017",   "A2017"),
    ("A5083",   "A5083"),
    ("A6061",   "A6061"),
    ("A7075",   "A7075"),
    ("A5056",   "A5056B"),
    ("A2011",   "A2011B"),
    ("SUS303",  "SUS303"),
    ("SUS316L", "SUS316L"),
    ("SUS316",  "SUS316"),
    ("SUS304",  "SUS304"),
    ("SUS440",  "SUS440C"),
    ("C1020",   "C1020P"),
    ("C1100",   "C1100P"),
    ("クロム",   "ｸﾛﾑｼﾞﾙｺﾆｳﾑ"),
    ("ｸﾛﾑ",    "ｸﾛﾑｼﾞﾙｺﾆｳﾑ"),
]

# 表示用の材質カテゴリ名（PDF用）
_SHEET_CATEGORY = {
    "A5052": "アルミニウム合金（板材）", "A2017": "アルミニウム合金（板材）",
    "A5083": "アルミニウム合金（板材）", "A6061": "アルミニウム合金（板材）",
    "A7075": "アルミニウム合金（板材）", "A5056B": "アルミニウム合金（棒材）",
    "A2011B": "アルミニウム合金（棒材）", "A2017B 75φ以下": "アルミニウム合金（棒材・75φ以下）",
    "A7075B 75φ以下": "アルミニウム合金（棒材・75φ以下）",
    "SUS303": "ステンレス鋼", "SUS304": "ステンレス鋼",
    "SUS316": "ステンレス鋼", "SUS316L": "ステンレス鋼", "SUS440C": "ステンレス鋼",
    "C1020P": "銅（無酸素銅）", "C1100P": "銅（タフピッチ銅）",
    "ｸﾛﾑｼﾞﾙｺﾆｳﾑ": "クロム銅・ジルコニウム銅",
}


def find_sheet_for_material(material_str: str) -> str | None:
    s = material_str.replace(" ", "").replace("　", "")
    for keyword, sheet in _MATERIAL_RULES:
        kw = keyword.replace(" ", "")
        if kw.upper() in s.upper() or kw in s:
            return sheet
    return None


def _excel_serial(d: date) -> int:
    from datetime import date as _d
    return (d - _d(1899, 12, 30)).days


def _make_doc_no(entry: dict) -> str:
    eid = entry.get("id", "")
    if len(eid) >= 14:
        return f"SXC-{eid[:8]}-{eid[8:14]}"
    return f"SXC-{eid}"


# ═══════════════════════════════════════════════════════════════════
# Excel 生成
#   テンプレートあり → xlrd+xlutils で .xls（シンクス正式フォーマット）
#   テンプレートなし → openpyxl で .xlsx（フォールバック）
# ═══════════════════════════════════════════════════════════════════

def _generate_xls_from_template(entry: dict) -> tuple[bytes | None, str | None]:
    """シンクス .xls テンプレートに顧客名・日付・品種を書き込んで返す"""
    try:
        import xlrd
        from xlutils.copy import copy as xl_copy
    except ImportError:
        return None, "xlrd/xlutils が未インストールです"

    tpl = _get_template()
    if not tpl:
        return None, "テンプレートなし"

    material = entry.get("material", "")
    sheet_name = find_sheet_for_material(material)
    if not sheet_name:
        return None, f"材質 '{material}' に対応するシートがありません"

    try:
        rb = xlrd.open_workbook(file_contents=tpl, formatting_info=True)
    except Exception as e:
        return None, f"テンプレート読み込みエラー: {e}"

    try:
        sheet_idx = rb.sheet_names().index(sheet_name)
    except ValueError:
        return None, f"シート '{sheet_name}' がテンプレートに見つかりません"

    wb = xl_copy(rb)
    ws = wb.get_sheet(sheet_idx)

    # 日付 (row1, col6)
    ws.write(1, 6, _excel_serial(date.today()))
    # 顧客名 (row2, col0)
    customer = (entry.get("customer") or "").strip()
    ws.write(2, 0, customer)
    # 品種・規格 (row19, col2)
    dims = (entry.get("dimensions") or "").strip()
    spec = material + (f"　{dims}" if dims else "")
    ws.write(19, 2, f"品種：{spec}")
    # 管理No. (row0, col7)
    eid = entry.get("id", "")
    doc_no = f"SXC-{eid[:8]}-{eid[8:14]}" if len(eid) >= 14 else f"SXC-{eid}"
    ws.write(0, 7, doc_no)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), None


def generate_xls(entry: dict) -> tuple[bytes | None, str | None]:
    """
    テンプレートがあれば xlrd+xlutils で .xls（シンクス正式フォーマット）を生成。
    テンプレートがなければ openpyxl で .xlsx（フォールバック）を生成。
    """
    # テンプレートあり → 正式フォーマット
    tpl_data, tpl_err = _generate_xls_from_template(entry)
    if tpl_data:
        return tpl_data, None
    # テンプレートなし → openpyxl フォールバック
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "openpyxl が未インストールです"

    material  = entry.get("material", "")
    customer  = (entry.get("customer") or "").strip()
    dims      = (entry.get("dimensions") or "").strip()
    purpose   = (entry.get("purpose") or "").strip()
    dest      = (entry.get("destination") or "").strip()
    overall   = entry.get("overall", "")
    steps     = entry.get("steps", [])
    rec       = entry.get("recommendation", "")

    sheet_name = find_sheet_for_material(material) or "（対応シートなし）"
    category   = _SHEET_CATEGORY.get(sheet_name, material)
    doc_no     = _make_doc_no(entry)
    today_str  = date.today().strftime("%Y/%m/%d")
    spec       = material + (f"　{dims}" if dims else "")
    target_item = _get_target_item(sheet_name)

    # ── スタイル定義 ──────────────────────────────────────────────
    thin   = Side(style="thin")
    med    = Side(style="medium")
    border_all  = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_med  = Border(top=med,  left=med,  right=med,  bottom=med)
    fill_header = PatternFill("solid", fgColor="CCCCCC")
    fill_step   = PatternFill("solid", fgColor="F0F0F0")
    fill_ok     = PatternFill("solid", fgColor="E8F5E9")
    fill_warn   = PatternFill("solid", fgColor="FFF8E1")
    fill_ng     = PatternFill("solid", fgColor="FFEBEE")
    OVERALL_FILL = {"OK": fill_ok, "要確認": fill_warn, "要許可": fill_ng}

    def cell_style(ws, row, col, value="", font_size=10, bold=False, wrap=False,
                   halign="left", valign="center", fill=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="游ゴシック", size=font_size, bold=bold)
        c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if fill:
            c.fill = fill
        if border:
            c.border = border
        return c

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 列幅
    col_widths = [6, 6, 18, 18, 28, 12, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 行1〜2: 管理No. / 発行日 ─────────────────────────────────
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 15
    cell_style(ws, 1, 7, "管理No.", bold=True, halign="right")
    cell_style(ws, 1, 8, doc_no)
    cell_style(ws, 2, 7, "発 行 日", bold=True, halign="right")
    cell_style(ws, 2, 8, today_str)

    # ── 行3: タイトル ────────────────────────────────────────────
    ws.row_dimensions[3].height = 24
    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1,
                value="輸出貿易管理令別表第一（1〜15項）　該非判定証明書")
    c.font = Font(name="游ゴシック", size=14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_med

    # ── 行4: 宛先 ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:C4")
    cell_style(ws, 4, 1, f"{customer}　御中", font_size=12, bold=True)

    # ── 行5〜7: 本文 + 発行者 ─────────────────────────────────────
    ws.row_dimensions[5].height = 14
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 14
    ws.merge_cells("A5:E5")
    ws.merge_cells("A6:E6")
    ws.merge_cells("A7:E7")
    cell_style(ws, 5, 1,
        "弊社取扱いの下記商品について、輸出貿易管理令別表第一／外国為替令別表の該当・非該当の別を下記のとおり証明いたします。",
        font_size=9, wrap=True)
    cell_style(ws, 6, 1, "なお、16項には「該当」します。", font_size=9)
    cell_style(ws, 7, 6, "神奈川県愛甲郡愛川町中津桜台4057-2", font_size=8, halign="right")
    ws.merge_cells("F7:H7")
    cell_style(ws, 5, 6, "株式会社シンクスコーポレーション", font_size=9, bold=True, halign="right")
    ws.merge_cells("F5:H5")
    cell_style(ws, 6, 6, "代表取締役　郡司　克彦", font_size=9, halign="right")
    ws.merge_cells("F6:H6")

    # ── 行9: 商品テーブルヘッダー ─────────────────────────────────
    ws.row_dimensions[9].height = 18
    HDR = ["No.", "商品等名称", "品種・規格・仕様・材質等",
           "形状", "対象項番", "判定結果", "備考"]
    merge_map = {2: "B9:C9"}  # 商品等名称は2列幅
    _h_cols = [1, 2, 4, 5, 6, 7, 8]
    _h_vals = HDR
    col_idxs = [1, 2, 4, 5, 6, 7, 8]
    for ci, val in zip(col_idxs, HDR):
        cell_style(ws, 9, ci, val, bold=True, halign="center",
                   fill=fill_header, border=border_all)
    ws.merge_cells("B9:C9")

    # ── 行10〜12: 商品データ行 ────────────────────────────────────
    ws.row_dimensions[10].height = 36
    ws.row_dimensions[11].height = 14
    ws.row_dimensions[12].height = 14

    of = OVERALL_FILL.get(overall)
    cell_style(ws, 10, 1, "1", halign="center", border=border_all)
    ws.merge_cells("B10:C10")
    cell_style(ws, 10, 2, category, wrap=True, valign="top", border=border_all)
    cell_style(ws, 10, 4, spec, wrap=True, valign="top", border=border_all)
    ws.merge_cells("E10:E12")
    cell_style(ws, 10, 5, target_item, halign="center", valign="center", border=border_all)
    ws.merge_cells("F10:F12")
    cell_style(ws, 10, 6, overall, bold=True, halign="center", valign="center",
               fill=of, border=border_all)
    ws.merge_cells("G10:H12")
    cell_style(ws, 10, 7, "AI一次判定", halign="center", border=border_all)

    # 形状・品種（縦）
    cell_style(ws, 11, 1, "", border=border_all)
    ws.merge_cells("B11:C11")
    cell_style(ws, 11, 2, f"  ・形状：{'板' if 'B' not in sheet_name else '棒'}", border=border_all)
    cell_style(ws, 11, 4, "", border=border_all)
    cell_style(ws, 12, 1, "", border=border_all)
    ws.merge_cells("B12:C12")
    cell_style(ws, 12, 2, f"  ・品種：{sheet_name}", border=border_all)
    cell_style(ws, 12, 4, "", border=border_all)

    # ── 行14〜15: 参考情報 ──────────────────────────────────────
    ws.row_dimensions[14].height = 14
    ws.row_dimensions[15].height = 14
    cell_style(ws, 14, 1, "【仕向け国】", bold=True)
    ws.merge_cells("B14:H14")
    cell_style(ws, 14, 2, dest)
    cell_style(ws, 15, 1, "【用　途】", bold=True)
    ws.merge_cells("B15:H15")
    cell_style(ws, 15, 2, purpose, wrap=True)

    # ── 行17: AIステップ表ヘッダー ────────────────────────────────
    ws.row_dimensions[17].height = 16
    ws.merge_cells("A17:H17")
    cell_style(ws, 17, 1, "■ AI一次判定 5ステップサマリー（参考情報）",
               bold=True, fill=PatternFill("solid", fgColor="E3F2FD"))

    ws.row_dimensions[18].height = 15
    step_hdrs = ["Step", "確認内容", "判定結果", "判定根拠（AI）", "信頼度"]
    step_cols  = [1, 2, 4, 5, 8]
    for ci, val in zip(step_cols, step_hdrs):
        cell_style(ws, 18, ci, val, bold=True, halign="center",
                   fill=fill_header, border=border_all)
    ws.merge_cells("B18:C18")
    ws.merge_cells("E18:G18")

    for i, s in enumerate(steps, start=19):
        ws.row_dimensions[i].height = 16
        res  = s.get("result", "")
        sfill = {"非該当": fill_ok, "ホワイト国": fill_ok, "問題なし": fill_ok,
                 "スキップ": fill_step}.get(res)
        if res in ("要確認",): sfill = fill_warn
        if res in ("該当", "非ホワイト国", "指定国", "輸出不可"): sfill = fill_ng
        cell_style(ws, i, 1, str(s.get("step", "")), halign="center",
                   fill=sfill, border=border_all)
        ws.merge_cells(f"B{i}:C{i}")
        cell_style(ws, i, 2, s.get("title", "")[:20], fill=sfill, border=border_all)
        cell_style(ws, i, 4, res, bold=True, halign="center", fill=sfill, border=border_all)
        ws.merge_cells(f"E{i}:G{i}")
        cell_style(ws, i, 5, s.get("reason", ""), wrap=True, fill=sfill, border=border_all)
        cell_style(ws, i, 8, s.get("confidence", ""), halign="center",
                   fill=sfill, border=border_all)

    # ── 推奨アクション ──────────────────────────────────────────
    R = 19 + len(steps) + 1
    ws.row_dimensions[R].height = 16
    ws.merge_cells(f"A{R}:H{R}")
    cell_style(ws, R, 1, f"【推奨アクション】 {rec}", bold=True,
               fill=PatternFill("solid", fgColor="FFF9C4"))

    # ── 注意書き ────────────────────────────────────────────────
    R2 = R + 2
    notices = [
        "※ 本証明書はAIによる一次スクリーニング結果に基づき発行しています。",
        "   最終的な該非判定は担当者が公式の「貨物・技術のマトリクス表」で確定してください。",
        "※ 本商品を加工して製品とする場合は、当該製品について別途判定が必要です。",
    ]
    for j, note in enumerate(notices):
        ws.merge_cells(f"A{R2+j}:H{R2+j}")
        nc = ws.cell(row=R2+j, column=1, value=note)
        nc.font = Font(name="游ゴシック", size=8, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), None


# ═══════════════════════════════════════════════════════════════════
# PDF 生成 (reportlab)
# ═══════════════════════════════════════════════════════════════════

def generate_pdf(entry: dict) -> tuple[bytes | None, str | None]:
    """該非判定証明書 PDF を生成する"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
    except ImportError:
        return None, "reportlab が未インストールです"

    # フォント登録
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
    FONT = "HeiseiKakuGo-W5"
    FONT_M = "HeiseiMin-W3"

    material  = entry.get("material", "")
    customer  = (entry.get("customer") or "").strip() or "（顧客名未入力）"
    dims      = (entry.get("dimensions") or "").strip()
    purpose   = (entry.get("purpose") or "").strip()
    dest      = (entry.get("destination") or "").strip()
    overall   = entry.get("overall", "")
    rec       = entry.get("recommendation", "")
    steps     = entry.get("steps", [])
    sheet_name = find_sheet_for_material(material) or "（対応シートなし）"
    category  = _SHEET_CATEGORY.get(sheet_name, material)
    doc_no    = _make_doc_no(entry)
    today_str = date.today().strftime("%Y年%m月%d日")
    spec      = material + (f"　{dims}" if dims else "")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4  # pt: 595.27 × 841.89

    # ── ヘルパー ──────────────────────────────────────────────────
    def font(size, bold=False):
        c.setFont(FONT, size)

    def text(x_mm, y_mm, s, size=9, align="left"):
        c.setFont(FONT, size)
        xp = x_mm * mm
        yp = H - y_mm * mm
        if align == "right":
            c.drawRightString(xp, yp, s)
        elif align == "center":
            c.drawCentredString(xp, yp, s)
        else:
            c.drawString(xp, yp, s)

    def hline(x1_mm, x2_mm, y_mm, width=0.5, dash=None):
        c.setLineWidth(width)
        if dash:
            c.setDash(*dash)
        c.line(x1_mm * mm, H - y_mm * mm, x2_mm * mm, H - y_mm * mm)
        if dash:
            c.setDash()

    def rect(x_mm, y_mm, w_mm, h_mm, fill=False, stroke=True):
        c.setLineWidth(0.5)
        c.rect(x_mm * mm, H - y_mm * mm - h_mm * mm, w_mm * mm, h_mm * mm,
               fill=1 if fill else 0, stroke=1 if stroke else 0)

    # ─────────────────────────────────────────────────────────────
    # ① ヘッダー部
    # ─────────────────────────────────────────────────────────────
    # タイトル
    text(105, 18, "輸出貿易管理令別表第一（1〜15項）", size=11, align="center")
    text(105, 25, "該  非  判  定  証  明  書", size=15, align="center")
    hline(15, 195, 28, width=1.5)

    # 右上：管理No. / 発行日
    text(195, 33, f"管理No.：{doc_no}", size=8, align="right")
    text(195, 38, f"発　行　日：{today_str}", size=8, align="right")

    # ─────────────────────────────────────────────────────────────
    # ② 宛先（顧客名）
    # ─────────────────────────────────────────────────────────────
    text(18, 44, customer, size=12)
    text(18 + len(customer) * 6 / mm + 3, 44, "　御中", size=11)
    hline(18, 100, 47, dash=[2, 2])

    # ─────────────────────────────────────────────────────────────
    # ③ 本文
    # ─────────────────────────────────────────────────────────────
    text(18, 54,
         "弊社取扱いの下記商品について、輸出貿易管理令別表第一／外国為替令別表の", size=9)
    text(18, 59,
         "該当・非該当の別を下記のとおり証明いたします。", size=9)
    text(18, 64,
         "なお、16項には「該当」します。", size=9)

    # ─────────────────────────────────────────────────────────────
    # ④ 発行者（右寄せ）
    # ─────────────────────────────────────────────────────────────
    text(195, 54, "神奈川県愛甲郡愛川町中津桜台4057-2", size=7.5, align="right")
    text(195, 59, "株式会社シンクスコーポレーション", size=9, align="right")
    text(195, 64, "代表取締役　郡司　克彦", size=9, align="right")

    # ─────────────────────────────────────────────────────────────
    # ⑤ 商品情報テーブル
    # ─────────────────────────────────────────────────────────────
    hline(15, 195, 70, width=1)
    Y_TBL = 71
    col_x   = [15, 18, 60, 130, 155, 175]  # mm  (left edges)
    col_w   = [3,  42, 70, 25,  20,  20]   # mm
    row_h   = 6                             # mm

    def tbl_row(y_mm, vals, sizes=None, bold_idx=()):
        """vals: list of str per column (6 cols)"""
        for i, v in enumerate(vals):
            cx = col_x[i]
            sz = (sizes[i] if sizes else 8)
            c.setFont(FONT, sz)
            c.drawString(cx * mm, H - (y_mm + 4) * mm, v)

    # ヘッダー行
    c.setFillColor(colors.HexColor("#EEEEEE"))
    rect(15, Y_TBL, 180, row_h, fill=True)
    c.setFillColor(colors.black)
    hdr = ["No", "商品等名称", "品種・規格・仕様・材質等", "対象項番", "判定結果", "備考"]
    tbl_row(Y_TBL, hdr, sizes=[7, 8, 8, 7, 8, 7])
    hline(15, 195, Y_TBL + row_h, width=0.8)

    # データ行
    Y_D = Y_TBL + row_h
    overall_color = {"OK": "#007700", "要確認": "#AA6600", "要許可": "#CC0000"}.get(overall, "#000000")
    rows_data = [
        ["1", category, spec, _get_target_item(sheet_name), overall, "AI一次判定"],
    ]
    for rd in rows_data:
        c.setFont(FONT, 8)
        for i, v in enumerate(rd):
            cx = col_x[i]
            if i == 4:  # 判定結果を色付け
                c.setFillColor(colors.HexColor(overall_color))
            c.drawString(cx * mm, H - (Y_D + 4) * mm, v)
            c.setFillColor(colors.black)
        hline(15, 195, Y_D + row_h, dash=[1, 1])
        Y_D += row_h

    hline(15, 195, Y_D, width=0.8)

    # ─────────────────────────────────────────────────────────────
    # ⑥ 参考情報（仕向け国・用途）
    # ─────────────────────────────────────────────────────────────
    Y_INFO = Y_D + 6
    text(15, Y_INFO,     f"【仕向け国】{dest}", size=8)
    text(15, Y_INFO + 6, f"【用  途】{purpose[:60]}{'…' if len(purpose) > 60 else ''}", size=8)

    # ─────────────────────────────────────────────────────────────
    # ⑦ AI判定結果サマリー（ステップ表）
    # ─────────────────────────────────────────────────────────────
    Y_STEP = Y_INFO + 16
    text(15, Y_STEP, "【AI一次判定 5ステップサマリー（参考）】", size=8)
    hline(15, 195, Y_STEP + 4, width=0.5)

    STEP_COLS = [15, 20, 55, 120, 145]  # mm
    STEP_HDR  = ["Step", "確認内容", "判定", "判定根拠（AI）", "信頼度"]
    STEP_W    = [5, 35, 65, 25, 15]

    Y_S = Y_STEP + 6
    c.setFillColor(colors.HexColor("#F5F5F5"))
    rect(15, Y_S, 180, 5, fill=True)
    c.setFillColor(colors.black)
    for i, h in enumerate(STEP_HDR):
        c.setFont(FONT, 7)
        c.drawString(STEP_COLS[i] * mm, H - (Y_S + 3.5) * mm, h)
    hline(15, 195, Y_S + 5, width=0.4)
    Y_S += 5

    RESULT_COLOR = {
        "非該当": "#007700", "ホワイト国": "#007700", "問題なし": "#007700",
        "スキップ": "#555555",
        "要確認": "#AA6600",
        "該当": "#CC0000", "非ホワイト国": "#CC0000", "指定国": "#CC0000", "輸出不可": "#CC0000",
    }

    for s in steps:
        res  = s.get("result", "")
        rsn  = s.get("reason", "")[:40]
        conf = s.get("confidence", "")
        rc   = RESULT_COLOR.get(res, "#000000")
        row_vals = [str(s.get("step", "")), s.get("title", "")[:18], res, rsn, conf]
        for i, v in enumerate(row_vals):
            c.setFont(FONT, 7.5)
            if i == 2:
                c.setFillColor(colors.HexColor(rc))
            c.drawString(STEP_COLS[i] * mm, H - (Y_S + 3.5) * mm, v)
            c.setFillColor(colors.black)
        hline(15, 195, Y_S + 5, dash=[1, 2])
        Y_S += 5

    hline(15, 195, Y_S, width=0.5)

    # ─────────────────────────────────────────────────────────────
    # ⑧ 推奨アクション
    # ─────────────────────────────────────────────────────────────
    Y_REC = Y_S + 8
    if rec:
        text(15, Y_REC, f"【推奨アクション】{rec[:80]}", size=8)
        Y_REC += 7

    # ─────────────────────────────────────────────────────────────
    # ⑨ 注意書き・フッター
    # ─────────────────────────────────────────────────────────────
    hline(15, 195, 270, width=0.5)
    NOTICE = [
        "※ 本証明書は AI による一次スクリーニング結果に基づき発行しています。",
        "   最終的な該非判定は担当者が公式の「貨物・技術のマトリクス表」で確定してください。",
        "※ 本商品を加工して製品とする場合は、当該製品について別途判定が必要です。",
        "   本証明書は弊社出荷時点の判定であり、加工後の製品の判定を保証するものではありません。",
    ]
    for i, n in enumerate(NOTICE):
        text(15, 274 + i * 5, n, size=7)

    c.save()
    buf.seek(0)
    return buf.read(), None


def _get_target_item(sheet_name: str) -> str:
    """材質シートに対応する輸出令別表第1の項番を返す（参考表示）"""
    MAP = {
        "A5052": "2(17)/5(5)", "A2017": "2(17)/5(5)", "A5083": "2(17)/5(5)",
        "A6061": "2(17)/5(5)", "A7075": "2(17)/5(5)",
        "A5056B": "2(17)/5(5)", "A2011B": "2(17)/5(5)",
        "A2017B 75φ以下": "2(17)/5(5)", "A7075B 75φ以下": "2(17)/5(5)",
        "SUS303": "5(5)", "SUS304": "5(5)", "SUS316": "5(5)",
        "SUS316L": "5(5)", "SUS440C": "5(5)",
        "C1020P": "なし", "C1100P": "なし", "ｸﾛﾑｼﾞﾙｺﾆｳﾑ": "5(5)",
    }
    return MAP.get(sheet_name, "要確認")
